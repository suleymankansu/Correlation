import pandas as pd
import tqdm
from sklearn.preprocessing import MinMaxScaler
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap, LinearSegmentedColormap
from matplotlib import cm
import numpy as np
import json


def get_coin_list():
    coin_list = []
    file1 = open('coinlist.txt', 'r')
    Lines = file1.readlines()

    for line in Lines:
        coin_list.append(line.strip())

    file1.close()
    return coin_list


def calculate_correlation(file):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Get the headers from the first row
        headers = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]

        # Create a DataFrame from the Excel data
        df = pd.DataFrame(sheet.iter_rows(values_only=True, min_col=2, max_col=sheet.max_column, min_row=2),
                          columns=headers)

        # Normalize the data using Min-Max scaling
        scaler = MinMaxScaler()
        df_normalized = pd.DataFrame(scaler.fit_transform(df), columns=df.columns)

        # Calculate the correlation matrix
        correlation_matrix = df_normalized.corr()

        return correlation_matrix

    except FileNotFoundError:
        print(f"File '{file}' not found.")
        return None


def average_correlation_matrices(*matrices):
    avg_matrix = pd.concat(matrices).groupby(level=0).mean()
    return avg_matrix


def write_correlation(filename, corr_matrix, min_corr):
    coin_list = get_coin_list()

    txtFile = open(filename, "w")
    output = ""
    if corr_matrix is not None:
        output += "Average Correlation Matrix:\n"
        for coin in coin_list:
            new_df = corr_matrix[corr_matrix[coin] > min_corr]
            new_df = new_df[new_df[coin] < 1]
            if not new_df.empty:
                output += "**************For coin " + coin + "*******************\n"
                output += str(new_df[coin]) + "\n"

    txtFile.write(output)
    txtFile.close()

    """
    fig = plt.figure(figsize=(50, 50))

    # 111: 1x1 grid, first subplot
    ax = fig.add_subplot(111)

    # normalize data using vmin, vmax
    cax = ax.matshow(close_correlation_matrix)

    # add a colorbar to a plot.


    # define ticks
    ticks = np.arange(0, 478, 1)

    # set x and y tick marks
    ax.set_xticks(ticks)
    ax.set_yticks(ticks)
    # set x and y tick labels
    ax.set_xticklabels(coin_list, rotation=90)
    ax.set_yticklabels(coin_list)

    top = cm.get_cmap('Oranges_r', 128)
    bottom = cm.get_cmap('Blues', 128)

    newcolors = np.vstack((top(np.linspace(0, 1, 19)),
                           bottom(np.linspace(1, 0, 1))))
    newcmp = ListedColormap(newcolors, name='OrangeBlue')
    viridis = cm.get_cmap('viridis', 12)
    cmap = ListedColormap([viridis, newcmp])
    psm = ax.pcolormesh(close_correlation_matrix, cmap=newcmp)
    fig.colorbar(psm)
    # draw a matrix using the correlations data
    plt.show()
    
    """
