from urllib.parse import urljoin
import requests
import openpyxl
import tqdm

BASE_URL = "https://api.binance.com"


def get_coin_list():
    f = open('data.json')

    coin_list = []
    file1 = open('coinlist.txt', 'r')
    Lines = file1.readlines()

    for line in Lines:
        coin_list.append(line.strip())

    f.close()
    return coin_list


def get_data(ticker, interval, limit):
    params = {
        'symbol': ticker,
        'interval': interval,
        'limit': limit
    }

    url = urljoin(BASE_URL, "/api/v3/klines")
    payload = {}
    headers = {
        'Content-Type': 'application/json',
    }

    response = requests.request("GET", url, headers=headers, params=params).json()
    return response


def save_to_excel(file, ticker, data, column_name):
    try:
        workbook = openpyxl.load_workbook(file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    next_column = len(sheet[1]) + 1

    sheet.cell(row=1, column=next_column, value=ticker)

    for index, candle in enumerate(data, start=2):
        price = float(candle[column_name])
        sheet.cell(row=index, column=next_column, value=price)

    # Save the workbook back to the file
    workbook.save(file)


def save_data():
    coin_list = get_coin_list()

    close_excel = "close4h.xlsx"
    high_excel = "high4h.xlsx"
    low_excel = "low4h.xlsx"

    for coin in tqdm.tqdm(range(len(coin_list))):
        data = get_data(coin_list[coin], "4h", "1000")
        save_to_excel(high_excel, coin_list[coin], data, column_name=2)
        save_to_excel(low_excel, coin_list[coin], data, column_name=3)
        save_to_excel(close_excel, coin_list[coin], data, column_name=4)


